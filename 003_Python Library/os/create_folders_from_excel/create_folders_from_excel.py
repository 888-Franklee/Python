import os
from openpyxl import load_workbook

excel_file = r"D:\Lee Encyclopaedia（李論百科全書）\004_Information Tech\Python\003_Python Pratice\os\create_folders_from_excel\names.xlsx"

wb = load_workbook(excel_file)
ws = wb.active  

base_folder = r"D:\Lee Encyclopaedia（李論百科全書）\004_Information Tech\Python\003_Python Pratice\os\create_folders_from_excel"
os.makedirs(base_folder, exist_ok=True)

# A2以降のセルの値を順に読み取ってフォルダ作成
for row in ws.iter_rows(min_row=2, max_col=1):
    name = row[0].value
    if name:  # 値がある場合のみ処理
        print(f"name: {name}")
        folder_path = os.path.join(base_folder, str(name))
        os.makedirs(folder_path, exist_ok=True)
        print(f"フォルダ作成: {folder_path}")
